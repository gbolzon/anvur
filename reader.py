import docx2txt
from unidecode import unidecode
import re
from openpyxl import load_workbook, Workbook
import sys

def longest(strlist):
    n=len(strlist)
    Max=0
    for i,thestr in enumerate(strlist):
        length = len(thestr)
        if length>Max:
            Max=length
            pos=i
    return strlist[pos]

def deblank(str):
    l =len(str)
    for i in range(l):
        if not str[i]==' ':
            break
    junk=str[i:]
    l=len(junk)
    for i in range(l):
        if not junk[l-i-1]==' ':
            break
    return junk[:l-i]

def clean_string(str):
    l =len(str)
    for i in range(l):
        if not str[i] in [" ", ".", "," , ";"] :
            break
    junk=str[i:]
    l=len(junk)
    for i in range(l):
        if not junk[l-i-1] in [" ", ".", "," , ";"] :
            break
    return junk[:l-i]

class paper():
    def __init__(self, authors,title,IF, review, year, completeline):
        self.authors=authors
        self.title = clean_string(title)
        self.IF = IF
        self.review = clean_string(review)
        self.year = year
        self.line = completeline
    def authorlist(self):
        return [p.surname for p in self.authors]
    def sectionlist(self):
        SECTIONS=set()
        for p in self.authors: SECTIONS.add(p.sez)
        return list(SECTIONS)


class person():
    def __init__(self, firstname, surname, sez):
        self.capitalfirstname= deblank(firstname)
        self.capitalsurname  = deblank(surname)
        self.sez             = deblank(sez)
        surnames = re.split("[\ \-]",self.capitalsurname)
        nSurnames = len(surnames)
        if nSurnames>1:
            L=surnames[0].lower().capitalize()
            r = re.search("[\ \-]",self.capitalsurname)
            sep = r.group()
            for surname in surnames[1:]:
                s = surname.lower().capitalize()
                L = L + sep + s
            self.surname = L
        else:
            self.surname=deblank(surname.lower().capitalize())

        names = self.capitalfirstname.rsplit(" ")
        L=""
        for name in names:
            s=name[0] + "."
            L=L + s 
        self.firstname = deblank(L)
        self.name = self.surname + ", " + self.firstname
        self.regex = self.surname + "[\ \,][\ ]*" + self.firstname[:2]
    def __eq__(self, person2):
        if (self.surname==person2.surname) & (self.firstname==person2.firstname):
            return True
        else:
            return False
        

def find_in_titles(TITLES,YEAR):
    '''Returns index'''
    for it, title in enumerate(TITLES): 
        if title.find(str(YEAR))>-1:
            break
    return it
# rm blanks, extra \n, moving 'in press', mv Pub 2017 in 2018, replace(.. -> .;) 
pub_doc_file="/Users/gbolzon/Documents/OGS/ANVUR/ELABORATI/PUBBLICAZIONI 2015-2019-1.docx"
Sez_xlsfile ="/Users/gbolzon/Documents/OGS/ANVUR/ELABORATI/personale I-III 01.11.2019_mp.xlsx"



A = docx2txt.process(pub_doc_file)
DECOD = unidecode(A)

regex="Publications \d\d\d\d"
TITLES = re.findall(regex,DECOD)
PUBS_STRING_BY_YEAR = re.split(regex,DECOD)[1:]


#Pers_xlsfile="/Users/gbolzon/Documents/OGS/ANVUR/ELABORATI/PersOGS I_III_2015_2019.xlsx"
#wb = load_workbook(filename=Pers_xlsfile, read_only=False,data_only=True)
#xls_sheet = str(YEAR)
#ws=wb[xls_sheet]

# OGS_LIST=[]
# for row in range(2,200): # authors at 01/01
#     if ws.cell(row=row, column=2).value is not None:
#         surname   = unidecode(ws.cell(row=row, column=2).value)
#         firstname = unidecode(ws.cell(row=row, column=3).value)
#         p = person(firstname, surname)
#         if p in OGS_LIST:
#             print "ERROR in xls pers file"
#             sys.exit()
#         OGS_LIST.append(p)
#
# for row in range(2,200):# authors at 12/31
#     if ws.cell(row=row, column=7).value is not None:
#         surname   = unidecode(ws.cell(row=row, column=7).value)
#         firstname = unidecode(ws.cell(row=row, column=8).value)
#         p = person(firstname, surname)
#         if not p in OGS_LIST: OGS_LIST.append(p)
#
wb = load_workbook(filename=Sez_xlsfile, read_only=False,data_only=True)
ws=wb['1 nov 2019']
OGS_SEZ_LIST=[]

for row in range(2,200):
    if ws.cell(row=row, column=3).value is not None:
        surname   = unidecode(ws.cell(row=row, column=3).value)
        firstname = unidecode(ws.cell(row=row, column=4).value)
        sez       = unidecode(ws.cell(row=row, column=5).value)
        p = person(firstname, surname, sez)
        if not p in OGS_SEZ_LIST: OGS_SEZ_LIST.append(p)


PAPERS=[]
for YEAR in range(2015, 2020):
    iYear = find_in_titles(TITLES, YEAR)
    YEAR_PUBS = PUBS_STRING_BY_YEAR[iYear]
    LINES = [ l for l in YEAR_PUBS.rsplit("\n") if len(l)> 1]

    title=''
    review_doi_IF=''
    year_string = "(%d)" %YEAR
    year_regex  = "\(%d\)" %YEAR
    
    for iline, line in enumerate(LINES):
        line = line.replace("\t","")
        line = line.replace("Monti-Birkenmeier","Monti")
        line = line.replace("Plasencia Linares", "Plasencia")
        line = line.replace("Melaku","")
        line = line.replace(" ,",",")
        if line == "": continue

        ind_year=line.find(year_string)
        if ind_year==-1:
            #print iline, line
            break

        #print iline, line[ind_year:ind_year+6]
        #continue
        del title
        del review_doi_IF
        CANDIDATE_AUTHORS=[]
        for p in OGS_SEZ_LIST:
            if line.find(p.surname) > -1:
                CANDIDATE_AUTHORS.append(p)
        #if len(CANDIDATE_AUTHORS)==0: print iline, " no AUTHORS"
    
        if line[:80].find(".") > -1:
            names_with_dot=True
            first_name_regex="[A-Z][\.][\,\;\ ]"
        else:
            names_with_dot=False
            first_name_regex="[A-Z][\,\ ]"

        Impact_factor_regex="IF[\ ]*[\=\:][\ ]*[0-9]*[\.\,][0-9]*"
        #first_name_regex="[A-Z][\.]*[\,\ ]"
        line_before_year, line_after_year= re.split(year_regex,line)
        strlist= re.split(first_name_regex, line_before_year)
        authors_first_name = re.findall(first_name_regex, line_before_year)
        authors_last_name = strlist[:-1]

        if  strlist[-1] == ' ':
            pos_end_of_authors = len(line_before_year)
        else:
            pos_end_of_authors = line.find(strlist[-1])
        all_but_authors=line[pos_end_of_authors:]
        #pos_end_of_authors= line.find(all_but_authors)
        authors_string=line[:pos_end_of_authors]
        AUTHORS=[]
        for a in CANDIDATE_AUTHORS:
            if len(re.findall(a.regex,authors_string)) > 0:
                AUTHORS.append(a)
            else:
                pass
                #print "rejecting " + a.name
                #print line
        #AUTHORS = [a for a in CANDIDATE_AUTHORS if len(re.findall(a.regex,authors_string)) > 0 ]
        #if len(AUTHORS)==0: print line

        try:
            title, review_doi_IF=re.split(year_regex,all_but_authors)
            #print title
        except:
            try:
                print "secondo tentativo"
                pos = all_but_authors[:10].find("201")
                if pos>-1:
                    title=longest(all_but_authors[pos+4:].rsplit(","))
                    review_doi_IF=""
            except:
                print "CANNOT find year in " + all_but_authors
        try:
            strIF = re.findall(Impact_factor_regex, all_but_authors)[0].replace(",",".").replace(":","=")
            pos_equal=strIF.find("=")
            IF = float(strIF[pos_equal+1:])
            review_doi = re.split(Impact_factor_regex, review_doi_IF)[0]
        except:
            IF = None
            #print "NO IMPACT FACTOR in line  " , iline
            review_doi = review_doi_IF

        if title=="":
            print "Please correct doc file in  line ", iline + 1
            sys.exit()
        A=paper(AUTHORS,title,IF, review_doi, YEAR,line)
        for p in PAPERS:
            if p.title == A.title:
                print "double in " + title
                sys.exit()
        PAPERS.append(A)
        if len(A.sectionlist()) >3 :
            print "4 sezioni"
            sys.exit()


#book = Workbook()
#sheet = book.active
book = load_workbook(filename="Anvur_pub.xlsx", read_only=False,data_only=False)
sheet=book['Sheet']


# SECTIONS E F G H=year I=IF J=Title K=Review M=auth1
for ip, p in enumerate(PAPERS):
    xls_row=ip+11
    strS1    = "E%d" %(xls_row)
    strS2    = "F%d" %(xls_row)
    strS3    = "G%d" %(xls_row)
    strIF    = "I%d" %(xls_row)
    strYear  = "H%d" %(xls_row)
    strTitle = "J%d" %(xls_row)
    str_Rev  = "K%d" %(xls_row)
    str_Line = "L%d" %(xls_row)
    str_bla  = "M%d" %(xls_row)
    sheet[strIF   ] = p.IF
    sheet[strYear ] = p.year
    sheet[strTitle] = p.title
    sheet[str_Rev ] = p.review
    sheet[str_Line] = p.line
    sheet[str_bla ] = " "
    #if p.title.startswith("Simulating the Effects of Alternative Managemen"):
    #    print "TROVATO a riga", xls_row
    if len(p.authors) > 0:
        SECTIONS = p.sectionlist()
        sheet[strS1] = SECTIONS[0]
        if len(SECTIONS)>=2: sheet[strS2] = SECTIONS[1]
        if len(SECTIONS)==3: sheet[strS3] = SECTIONS[2]

        for ia, a in enumerate(p.authors):
            str_author="%s%d" %(chr(ia+78),xls_row)
            sheet[str_author] = a.surname


book.save("sample.xlsx")

sys.exit()

OUT_LINES=[]
for ip, p in enumerate(PAPERS):
    authors_string=''
    for a in p.authors: authors_string+=a.surname + "\t"

    outline= "%s\t%d\t%s\t%s\t \t%s\n"  %(p.IF , YEAR, p.title , p.review, authors_string)
    OUT_LINES.append(outline )

fid=open("2019.txt",'wt')
fid.writelines(OUT_LINES)
fid.close()

